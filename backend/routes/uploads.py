import os
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import User, WireMapping, db
from extensions import db as _db
from openpyxl import load_workbook

upload_bp = Blueprint("uploads", __name__)


@upload_bp.route("/wire-mapping", methods=["POST"])
@jwt_required()
def upload_wire_mapping():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({"error": "Admin access required"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only Excel files (.xlsx, .xls) are supported"}), 400

    mode = request.form.get("mode", "replace")  # replace or append

    try:
        wb = load_workbook(file)
        ws = wb.active

        if mode == "replace":
            WireMapping.query.delete()
            db.session.commit()

        added = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                skipped += 1
                continue

            fil_val = str(row[0]).strip() if row[0] else ""
            ccfe_val = str(row[1]).strip() if row[1] else ""

            if not fil_val or not ccfe_val:
                skipped += 1
                continue

            existing = WireMapping.query.filter_by(fil=fil_val, ccfe=ccfe_val).first()
            if existing:
                skipped += 1
                continue

            mapping = WireMapping(fil=fil_val, ccfe=ccfe_val)
            db.session.add(mapping)
            added += 1

        db.session.commit()

        total_mappings = WireMapping.query.count()
        return jsonify({
            "message": f"Upload complete: {added} added, {skipped} skipped",
            "added": added,
            "skipped": skipped,
            "total_mappings": total_mappings,
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to process file: {str(e)}"}), 500


@upload_bp.route("/wire-mapping", methods=["GET"])
@jwt_required()
def list_wire_mappings():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({"error": "Admin access required"}), 403

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    search = request.args.get("search", "")

    query = WireMapping.query
    if search:
        query = query.filter(
            db.or_(
                WireMapping.fil.ilike(f"%{search}%"),
                WireMapping.ccfe.ilike(f"%{search}%"),
            )
        )

    pagination = query.order_by(WireMapping.fil).paginate(page=page, per_page=per_page, error_out=False)
    return jsonify({
        "items": [w.to_dict() for w in pagination.items],
        "total": pagination.total,
        "page": pagination.page,
        "per_page": pagination.per_page,
        "pages": pagination.pages,
    }), 200


@upload_bp.route("/wire-mapping/<int:mapping_id>", methods=["DELETE"])
@jwt_required()
def delete_wire_mapping(mapping_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({"error": "Admin access required"}), 403

    mapping = WireMapping.query.get_or_404(mapping_id)
    db.session.delete(mapping)
    db.session.commit()
    return jsonify({"message": "Wire mapping deleted"}), 200


@upload_bp.route("/wire-mapping/clear", methods=["POST"])
@jwt_required()
def clear_wire_mappings():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({"error": "Admin access required"}), 403

    count = WireMapping.query.count()
    WireMapping.query.delete()
    db.session.commit()
    return jsonify({"message": f"Deleted {count} wire mappings"}), 200
