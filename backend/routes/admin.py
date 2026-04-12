from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import (
    User, QRCode, Area, Poste, TypeScrap, Raison,
    WireMapping, ScrapSession, ScrapEntry, OperatorCode, db,
)
from sqlalchemy import func, extract
import io

admin_bp = Blueprint("admin", __name__)


def _admin_required():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return False
    return user


# ── Users CRUD ──────────────────────────────────────────────────────────────
@admin_bp.route("/users", methods=["GET"])
@jwt_required()
def list_users():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    users = User.query.order_by(User.created_at.desc()).all()
    return jsonify([u.to_dict() for u in users]), 200


@admin_bp.route("/users", methods=["POST"])
@jwt_required()
def create_user():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "")
    email = data.get("email", "").strip() or None
    full_name = data.get("full_name", "").strip() or None

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({"error": "Username already exists"}), 400

    user = User(username=username, email=email, full_name=full_name, is_admin=True)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    return jsonify(user.to_dict()), 201


@admin_bp.route("/users/<int:user_id>", methods=["PUT"])
@jwt_required()
def update_user(user_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    user = User.query.get_or_404(user_id)
    data = request.get_json()

    if "email" in data:
        user.email = data["email"].strip() or None
    if "full_name" in data:
        user.full_name = data["full_name"].strip() or None
    if "is_admin" in data:
        user.is_admin = data["is_admin"]
    if "password" in data and data["password"]:
        user.set_password(data["password"])

    db.session.commit()
    return jsonify(user.to_dict()), 200


@admin_bp.route("/users/<int:user_id>", methods=["DELETE"])
@jwt_required()
def delete_user(user_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    if admin.id == user_id:
        return jsonify({"error": "Cannot delete your own account"}), 400

    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    return jsonify({"message": "User deleted"}), 200


# ── Scrap Sessions CRUD ─────────────────────────────────────────────────────
@admin_bp.route("/sessions", methods=["GET"])
@jwt_required()
def list_sessions():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    status = request.args.get("status")
    segment = request.args.get("segment")
    equipe = request.args.get("equipe")
    ligne = request.args.get("ligne")
    semaine = request.args.get("semaine", type=int)
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    query = ScrapSession.query

    if status:
        query = query.filter(ScrapSession.status == status)
    if segment:
        query = query.filter(ScrapSession.segment.ilike(f"%{segment}%"))
    if equipe:
        query = query.filter(ScrapSession.equipe.ilike(f"%{equipe}%"))
    if ligne:
        query = query.filter(ScrapSession.ligne.ilike(f"%{ligne}%"))
    if semaine:
        query = query.filter(ScrapSession.semaine == semaine)
    if date_from:
        query = query.filter(ScrapSession.date >= date_from)
    if date_to:
        query = query.filter(ScrapSession.date <= date_to)

    query = query.order_by(ScrapSession.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        "items": [s.to_dict(include_entries=True) for s in pagination.items],
        "total": pagination.total,
        "page": pagination.page,
        "per_page": pagination.per_page,
        "pages": pagination.pages,
    }), 200


@admin_bp.route("/sessions/<int:session_id>", methods=["GET"])
@jwt_required()
def get_session(session_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    session = ScrapSession.query.get_or_404(session_id)
    return jsonify(session.to_dict(include_entries=True)), 200


@admin_bp.route("/sessions/<int:session_id>", methods=["PUT"])
@jwt_required()
def update_session(session_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    session = ScrapSession.query.get_or_404(session_id)
    data = request.get_json()

    if "total_weight" in data:
        session.total_weight = float(data["total_weight"])
    if "status" in data:
        session.status = data["status"]
    if "semaine" in data:
        session.semaine = data["semaine"]

    db.session.commit()
    return jsonify(session.to_dict()), 200


@admin_bp.route("/sessions/<int:session_id>", methods=["DELETE"])
@jwt_required()
def delete_session(session_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    session = ScrapSession.query.get_or_404(session_id)
    db.session.delete(session)
    db.session.commit()
    return jsonify({"message": "Session deleted"}), 200


# ── Scrap Entries CRUD ──────────────────────────────────────────────────────
@admin_bp.route("/entries", methods=["GET"])
@jwt_required()
def list_entries():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)

    query = ScrapEntry.query.order_by(ScrapEntry.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        "items": [e.to_dict() for e in pagination.items],
        "total": pagination.total,
        "page": pagination.page,
        "per_page": pagination.per_page,
        "pages": pagination.pages,
    }), 200


@admin_bp.route("/entries/<int:entry_id>", methods=["PUT"])
@jwt_required()
def update_entry(entry_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    entry = ScrapEntry.query.get_or_404(entry_id)
    data = request.get_json()

    if "area_id" in data:
        entry.area_id = data["area_id"]
    if "type_scrap_id" in data:
        entry.type_scrap_id = data["type_scrap_id"]
    if "poste_id" in data:
        entry.poste_id = data["poste_id"]
    if "raison_id" in data:
        entry.raison_id = data["raison_id"]
    if "numero_piece" in data:
        entry.numero_piece = data["numero_piece"]
    if "fil" in data:
        entry.fil = data["fil"]
    if "quantite" in data:
        entry.quantite = int(data["quantite"])

    db.session.commit()
    return jsonify(entry.to_dict()), 200


@admin_bp.route("/entries/<int:entry_id>", methods=["DELETE"])
@jwt_required()
def delete_entry(entry_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    entry = ScrapEntry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    return jsonify({"message": "Entry deleted"}), 200


# ── Dashboard / Statistics ──────────────────────────────────────────────────
@admin_bp.route("/dashboard", methods=["GET"])
@jwt_required()
def dashboard():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    # Clean up orphaned pending sessions with 0 entries
    orphaned = ScrapSession.query.filter_by(status="pending_weight").all()
    for s in orphaned:
        if s.entries.count() == 0:
            db.session.delete(s)
    db.session.commit()

    # Total counts
    total_sessions = ScrapSession.query.count()
    completed_sessions = ScrapSession.query.filter_by(status="completed").count()
    pending_sessions = ScrapSession.query.filter_by(status="pending_weight").count()
    total_entries = ScrapEntry.query.count()
    total_weight = db.session.query(func.coalesce(func.sum(ScrapSession.total_weight), 0)).filter(
        ScrapSession.status == "completed"
    ).scalar()
    total_quantite = db.session.query(func.coalesce(func.sum(ScrapEntry.quantite), 0)).scalar()

    # Scrap by type
    scrap_by_type = (
        db.session.query(TypeScrap.name, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
        .join(ScrapEntry, ScrapEntry.type_scrap_id == TypeScrap.id)
        .group_by(TypeScrap.name)
        .all()
    )

    # Scrap by area
    scrap_by_area = (
        db.session.query(Area.name, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
        .join(ScrapEntry, ScrapEntry.area_id == Area.id)
        .group_by(Area.name)
        .all()
    )

    # Scrap by segment
    scrap_by_segment = (
        db.session.query(ScrapSession.segment, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
        .join(ScrapEntry, ScrapEntry.session_id == ScrapSession.id)
        .group_by(ScrapSession.segment)
        .all()
    )

    # Scrap by equipe
    scrap_by_equipe = (
        db.session.query(ScrapSession.equipe, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
        .join(ScrapEntry, ScrapEntry.session_id == ScrapSession.id)
        .group_by(ScrapSession.equipe)
        .all()
    )

    # Scrap by raison
    scrap_by_raison = (
        db.session.query(Raison.name, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
        .join(ScrapEntry, ScrapEntry.raison_id == Raison.id)
        .group_by(Raison.name)
        .all()
    )

    # Weekly trend (last 12 weeks)
    weekly_trend = (
        db.session.query(
            ScrapSession.semaine,
            func.count(ScrapEntry.id),
            func.coalesce(func.sum(ScrapEntry.quantite), 0),
            func.coalesce(func.sum(ScrapSession.total_weight), 0),
        )
        .join(ScrapEntry, ScrapEntry.session_id == ScrapSession.id)
        .group_by(ScrapSession.semaine)
        .order_by(ScrapSession.semaine.desc())
        .limit(12)
        .all()
    )

    # Daily trend (last 30 days)
    daily_trend = (
        db.session.query(
            ScrapSession.date,
            func.count(ScrapEntry.id),
            func.coalesce(func.sum(ScrapEntry.quantite), 0),
        )
        .join(ScrapEntry, ScrapEntry.session_id == ScrapSession.id)
        .group_by(ScrapSession.date)
        .order_by(ScrapSession.date.desc())
        .limit(30)
        .all()
    )

    # Top postes by scrap count
    top_postes = (
        db.session.query(Poste.name, Area.name, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
        .join(Area, Poste.area_id == Area.id)
        .join(ScrapEntry, ScrapEntry.poste_id == Poste.id)
        .group_by(Poste.name, Area.name)
        .order_by(func.count(ScrapEntry.id).desc())
        .limit(10)
        .all()
    )

    # Top fils by scrap count
    top_fils = (
        db.session.query(ScrapEntry.fil, ScrapEntry.numero_piece, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
        .group_by(ScrapEntry.fil, ScrapEntry.numero_piece)
        .order_by(func.count(ScrapEntry.id).desc())
        .limit(10)
        .all()
    )

    return jsonify({
        "summary": {
            "total_sessions": total_sessions,
            "completed_sessions": completed_sessions,
            "pending_sessions": pending_sessions,
            "total_entries": total_entries,
            "total_weight": float(total_weight),
            "total_quantite": int(total_quantite),
        },
        "scrap_by_type": [{"name": r[0], "count": r[1], "quantite": int(r[2])} for r in scrap_by_type],
        "scrap_by_area": [{"name": r[0], "count": r[1], "quantite": int(r[2])} for r in scrap_by_area],
        "scrap_by_segment": [{"name": r[0], "count": r[1], "quantite": int(r[2])} for r in scrap_by_segment],
        "scrap_by_equipe": [{"name": r[0], "count": r[1], "quantite": int(r[2])} for r in scrap_by_equipe],
        "scrap_by_raison": [{"name": r[0], "count": r[1], "quantite": int(r[2])} for r in scrap_by_raison],
        "weekly_trend": [{"semaine": r[0], "count": r[1], "quantite": int(r[2]), "weight": float(r[3])} for r in weekly_trend],
        "daily_trend": [{"date": r[0].isoformat() if r[0] else None, "count": r[1], "quantite": int(r[2])} for r in daily_trend],
        "top_postes": [{"poste": r[0], "area": r[1], "count": r[2], "quantite": int(r[3])} for r in top_postes],
        "top_fils": [{"fil": r[0], "ccfe": r[1], "count": r[2], "quantite": int(r[3])} for r in top_fils],
    }), 200


# ── Export ──────────────────────────────────────────────────────────────────
@admin_bp.route("/export", methods=["GET"])
@jwt_required()
def export_data():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    format_type = request.args.get("format", "csv")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    segment = request.args.get("segment")
    equipe = request.args.get("equipe")
    ligne = request.args.get("ligne")
    semaine = request.args.get("semaine", type=int)

    query = (
        db.session.query(ScrapEntry, ScrapSession)
        .join(ScrapSession, ScrapEntry.session_id == ScrapSession.id)
    )

    if date_from:
        query = query.filter(ScrapSession.date >= date_from)
    if date_to:
        query = query.filter(ScrapSession.date <= date_to)
    if segment:
        query = query.filter(ScrapSession.segment.ilike(f"%{segment}%"))
    if equipe:
        query = query.filter(ScrapSession.equipe.ilike(f"%{equipe}%"))
    if ligne:
        query = query.filter(ScrapSession.ligne.ilike(f"%{ligne}%"))
    if semaine:
        query = query.filter(ScrapSession.semaine == semaine)

    results = query.order_by(ScrapSession.date.desc(), ScrapSession.id, ScrapEntry.id).all()

    # Track which sessions have had their weight already written
    weight_written_sessions = set()

    if format_type == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Semaine", "Date", "Segment", "Equipe", "Ligne",
            "Area", "Type Scrap", "Numéro de pièce", "Fil",
            "Quantité", "Raison", "Poste", "Poids Total (kg)", "Statut"
        ])

        for entry, session in results:
            area = Area.query.get(entry.area_id)
            type_scrap = TypeScrap.query.get(entry.type_scrap_id)
            raison = Raison.query.get(entry.raison_id)
            poste = Poste.query.get(entry.poste_id)

            # Show total weight only on the first row of each session
            if session.id not in weight_written_sessions:
                weight_written_sessions.add(session.id)
                weight_val = session.total_weight if session.total_weight else ""
            else:
                weight_val = ""

            writer.writerow([
                session.semaine,
                session.date.isoformat() if session.date else "",
                session.segment,
                session.equipe,
                session.ligne,
                area.name if area else "",
                type_scrap.name if type_scrap else "",
                entry.numero_piece,
                entry.fil,
                entry.quantite,
                raison.name if raison else "",
                poste.name if poste else "",
                weight_val,
                session.status,
            ])

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"scrap_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )

    elif format_type == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Colors & Styles ──────────────────────────────────────────────────
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        DATA_FONT = Font(name="Calibri", size=10)
        THIN_BORDER = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        ALT_ROW_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
        TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2E86C1")

        def style_header_row(ws, num_cols):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER

        def style_data_cell(ws, row, col, alt=False):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if alt:
                cell.fill = ALT_ROW_FILL

        def auto_width(ws, num_cols, min_w=10, max_w=30):
            for col in range(1, num_cols + 1):
                max_len = min_w
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, min(len(str(cell.value)) + 2, max_w))
                ws.column_dimensions[get_column_letter(col)].width = max_len

        # ── Sheet 1: Scrap Data (no Poids Total) ────────────────────────────
        ws1 = wb.active
        ws1.title = "Données Scrap"
        headers1 = [
            "Semaine", "Date", "Segment", "Équipe", "Ligne",
            "Area", "Type Scrap", "N° Pièce / CCFE", "Fil",
            "Quantité", "Raison", "Poste"
        ]
        ws1.append(headers1)
        style_header_row(ws1, len(headers1))

        for idx, (entry, session) in enumerate(results):
            area = Area.query.get(entry.area_id)
            type_scrap = TypeScrap.query.get(entry.type_scrap_id)
            raison = Raison.query.get(entry.raison_id)
            poste = Poste.query.get(entry.poste_id)

            row_num = idx + 2
            values = [
                session.semaine,
                session.date.isoformat() if session.date else "",
                session.segment,
                session.equipe,
                session.ligne,
                area.name if area else "",
                type_scrap.name if type_scrap else "",
                entry.numero_piece or "",
                entry.fil,
                entry.quantite,
                raison.name if raison else "",
                poste.name if poste else "",
            ]
            ws1.append(values)
            for col in range(1, len(values) + 1):
                style_data_cell(ws1, row_num, col, alt=(idx % 2 == 1))

        auto_width(ws1, len(headers1))
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{ws1.max_row}"
        ws1.freeze_panes = "A2"

        # ── Sheet 2: Résumé Poids ───────────────────────────────────────────
        ws2 = wb.create_sheet("Résumé Poids")
        headers2 = ["Semaine", "Date", "Segment", "Équipe", "Ligne", "Type Scrap", "Poids Total (kg)"]
        ws2.append(headers2)
        style_header_row(ws2, len(headers2))

        # Group sessions by (session, type_scrap) for weight summary
        seen_sessions = {}
        for entry, session in results:
            type_scrap = TypeScrap.query.get(entry.type_scrap_id)
            key = (session.id, entry.type_scrap_id)
            if key not in seen_sessions:
                seen_sessions[key] = {
                    "semaine": session.semaine,
                    "date": session.date.isoformat() if session.date else "",
                    "segment": session.segment,
                    "equipe": session.equipe,
                    "ligne": session.ligne,
                    "type_scrap": type_scrap.name if type_scrap else "",
                    "weight": session.total_weight or 0,
                }

        for idx, row_data in enumerate(seen_sessions.values()):
            row_num = idx + 2
            values = [
                row_data["semaine"],
                row_data["date"],
                row_data["segment"],
                row_data["equipe"],
                row_data["ligne"],
                row_data["type_scrap"],
                row_data["weight"],
            ]
            ws2.append(values)
            for col in range(1, len(values) + 1):
                style_data_cell(ws2, row_num, col, alt=(idx % 2 == 1))
            # Highlight weight column
            weight_cell = ws2.cell(row=row_num, column=7)
            weight_cell.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")

        auto_width(ws2, len(headers2))
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{ws2.max_row}"
        ws2.freeze_panes = "A2"

        # ── Sheet 3: Statistiques & Graphiques ──────────────────────────────
        ws3 = wb.create_sheet("Statistiques")

        # Title
        ws3.merge_cells("A1:F1")
        title_cell = ws3.cell(row=1, column=1, value="Statistiques Scrap")
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 35

        # -- Table: Scrap par Type --
        ws3.cell(row=3, column=1, value="Scrap par Type").font = SUBTITLE_FONT
        type_headers = ["Type Scrap", "Nombre d'entrées", "Quantité totale"]
        for ci, h in enumerate(type_headers, 1):
            c = ws3.cell(row=4, column=ci, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

        scrap_by_type_data = (
            db.session.query(TypeScrap.name, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
            .join(ScrapEntry, ScrapEntry.type_scrap_id == TypeScrap.id)
            .group_by(TypeScrap.name)
            .all()
        )
        for ri, (name, cnt, qty) in enumerate(scrap_by_type_data):
            ws3.cell(row=5 + ri, column=1, value=name).border = THIN_BORDER
            ws3.cell(row=5 + ri, column=2, value=cnt).border = THIN_BORDER
            ws3.cell(row=5 + ri, column=3, value=int(qty)).border = THIN_BORDER

        type_data_end = 5 + len(scrap_by_type_data)

        # Bar chart: Scrap par Type
        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Scrap par Type"
        chart1.y_axis.title = "Quantité"
        chart1.x_axis.title = "Type"
        chart1.style = 10
        chart1.width = 18
        chart1.height = 12
        data_ref = Reference(ws3, min_col=3, min_row=4, max_row=type_data_end - 1)
        cats_ref = Reference(ws3, min_col=1, min_row=5, max_row=type_data_end - 1)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        chart1.shape = 4
        ws3.add_chart(chart1, "E3")

        # -- Table: Scrap par Area --
        area_start_row = type_data_end + 2
        ws3.cell(row=area_start_row, column=1, value="Scrap par Area").font = SUBTITLE_FONT
        for ci, h in enumerate(["Area", "Nombre d'entrées", "Quantité totale"], 1):
            c = ws3.cell(row=area_start_row + 1, column=ci, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

        scrap_by_area_data = (
            db.session.query(Area.name, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
            .join(ScrapEntry, ScrapEntry.area_id == Area.id)
            .group_by(Area.name)
            .all()
        )
        for ri, (name, cnt, qty) in enumerate(scrap_by_area_data):
            ws3.cell(row=area_start_row + 2 + ri, column=1, value=name).border = THIN_BORDER
            ws3.cell(row=area_start_row + 2 + ri, column=2, value=cnt).border = THIN_BORDER
            ws3.cell(row=area_start_row + 2 + ri, column=3, value=int(qty)).border = THIN_BORDER

        area_data_end = area_start_row + 2 + len(scrap_by_area_data)

        # Pie chart: Scrap par Area
        chart2 = PieChart()
        chart2.title = "Répartition par Area"
        chart2.style = 10
        chart2.width = 18
        chart2.height = 12
        data_ref2 = Reference(ws3, min_col=3, min_row=area_start_row + 1, max_row=area_data_end - 1)
        cats_ref2 = Reference(ws3, min_col=1, min_row=area_start_row + 2, max_row=area_data_end - 1)
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats_ref2)
        ws3.add_chart(chart2, "E" + str(area_start_row))

        # -- Table: Scrap par Raison --
        raison_start_row = area_data_end + 2
        ws3.cell(row=raison_start_row, column=1, value="Scrap par Raison").font = SUBTITLE_FONT
        for ci, h in enumerate(["Raison", "Nombre d'entrées", "Quantité totale"], 1):
            c = ws3.cell(row=raison_start_row + 1, column=ci, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

        scrap_by_raison_data = (
            db.session.query(Raison.name, func.count(ScrapEntry.id), func.coalesce(func.sum(ScrapEntry.quantite), 0))
            .join(ScrapEntry, ScrapEntry.raison_id == Raison.id)
            .group_by(Raison.name)
            .all()
        )
        for ri, (name, cnt, qty) in enumerate(scrap_by_raison_data):
            ws3.cell(row=raison_start_row + 2 + ri, column=1, value=name).border = THIN_BORDER
            ws3.cell(row=raison_start_row + 2 + ri, column=2, value=cnt).border = THIN_BORDER
            ws3.cell(row=raison_start_row + 2 + ri, column=3, value=int(qty)).border = THIN_BORDER

        raison_data_end = raison_start_row + 2 + len(scrap_by_raison_data)

        # Bar chart: Scrap par Raison
        chart3 = BarChart()
        chart3.type = "col"
        chart3.title = "Scrap par Raison"
        chart3.y_axis.title = "Quantité"
        chart3.style = 10
        chart3.width = 18
        chart3.height = 12
        data_ref3 = Reference(ws3, min_col=3, min_row=raison_start_row + 1, max_row=raison_data_end - 1)
        cats_ref3 = Reference(ws3, min_col=1, min_row=raison_start_row + 2, max_row=raison_data_end - 1)
        chart3.add_data(data_ref3, titles_from_data=True)
        chart3.set_categories(cats_ref3)
        ws3.add_chart(chart3, "E" + str(raison_start_row))

        auto_width(ws3, 6)

        # ── Save ─────────────────────────────────────────────────────────────
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)

        return send_file(
            xlsx_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"scrap_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

    return jsonify({"error": "Unsupported format. Use csv or xlsx"}), 400


# ── Operator Codes CRUD ───────────────────────────────────────────────────
@admin_bp.route("/operator-codes", methods=["GET"])
@jwt_required()
def list_operator_codes():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    codes = OperatorCode.query.order_by(OperatorCode.created_at.desc()).all()
    return jsonify([c.to_dict() for c in codes]), 200


@admin_bp.route("/operator-codes", methods=["POST"])
@jwt_required()
def create_operator_code():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    data = request.get_json()
    code = (data.get("code") or "").strip()
    label = (data.get("label") or "").strip()

    if not code:
        return jsonify({"error": "Code is required"}), 400

    if OperatorCode.query.filter_by(code=code).first():
        return jsonify({"error": "Code already exists"}), 400

    op_code = OperatorCode(code=code, label=label or "Code opérateur")
    db.session.add(op_code)
    db.session.commit()
    return jsonify(op_code.to_dict()), 201


@admin_bp.route("/operator-codes/<int:code_id>", methods=["PUT"])
@jwt_required()
def update_operator_code(code_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    op_code = OperatorCode.query.get_or_404(code_id)
    data = request.get_json()

    if "code" in data:
        new_code = data["code"].strip()
        if new_code != op_code.code and OperatorCode.query.filter_by(code=new_code).first():
            return jsonify({"error": "Code already exists"}), 400
        op_code.code = new_code
    if "label" in data:
        op_code.label = data["label"].strip() or "Code opérateur"
    if "is_active" in data:
        op_code.is_active = data["is_active"]

    db.session.commit()
    return jsonify(op_code.to_dict()), 200


@admin_bp.route("/operator-codes/<int:code_id>", methods=["DELETE"])
@jwt_required()
def delete_operator_code(code_id):
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    op_code = OperatorCode.query.get_or_404(code_id)
    db.session.delete(op_code)
    db.session.commit()
    return jsonify({"message": "Operator code deleted"}), 200


@admin_bp.route("/operator-codes/generate", methods=["POST"])
@jwt_required()
def generate_operator_code():
    admin = _admin_required()
    if not admin:
        return jsonify({"error": "Admin access required"}), 403

    import random, string
    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    while OperatorCode.query.filter_by(code=code).first():
        code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

    data = request.get_json() or {}
    label = (data.get("label") or "").strip() or "Code opérateur"

    op_code = OperatorCode(code=code, label=label)
    db.session.add(op_code)
    db.session.commit()
    return jsonify(op_code.to_dict()), 201
